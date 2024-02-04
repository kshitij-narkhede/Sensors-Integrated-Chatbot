from time import *
import os
import subprocess,sys
import openpyxl
from time import *
import os
from subprocess import call

while True:
    os.system('start cmd /k start1\\stress4.cmd')
    sleep(5)
    wb_obj = openpyxl.load_workbook(r'Sensors_Data\Book1.xlsx')
    sheet_obj = wb_obj.active
    sum1=0
    for i in range(8,38):
         cell_obj = sheet_obj.cell(row = i, column = 2)
         sum1=sum1+cell_obj.value
    average1=int(sum1/30)
    print(average1)
    bpm_range=105
    sum2=0
    for i in range(8,38):
        cell_obj = sheet_obj.cell(row = i, column = 3)
        sum2=sum2+cell_obj.value
    average2=int(sum2/30)
    print(average2)
    bpm_range_lower=60
    bpm_range_higher=105
    gsr_range_lower=120
    gsr_range_higher=160
    
    #if (average1<bpm_range_lower &average1>bpm_range_higher & average2<gsr_range_lower &average2>gsr_range_higher):
    if True: 
        call(["python", "kvn.py"])
        break
    
    else:
        print("You are not in stress!! Still Deactivated")
    p = subprocess.Popen(["powershell.exe", "D:\SIP_Team_C\\start1\\stress6.ps1"], stdout=sys.stdout)
    p.communicate()
    sleep(5)



    # os.system('start cmd /k start1\\stress1.cmd')
    # sleep(5)
    # os.system('start cmd /k start1\\stress2.cmd')
    # sleep(15)
    # os.system('start cmd /k start1\\stress3.cmd')
    
# else:
#     print("You are not in stress!! Still Deactivated")
#i am kshitij Hello
#120 to 135 gsr
#105 hrv
#137  149 

